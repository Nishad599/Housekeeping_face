import os

target_file = r"d:\projects\housekeeping_attendancce\face-attendance\app\services\attendance_service.py"
func_code = """

def generate_individual_report_excel(staff, records, year, month):
    \"\"\"Generate an Excel workbook for a single staff member's monthly attendance.\"\"\"
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime
    
    month_label = datetime(year, month, 1).strftime("%B %Y")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = staff.employee_id[:31]  # Excel tab limit

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    sum_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    status_colors = {"Present": "D5F5E3", "Absent": "FADBD8",
                     "Partial": "FEF9E7", "Weekly Off": "EBEDEF"}
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"MONTHLY ATTENDANCE — {month_label.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"Employee: {staff.name}  ({staff.employee_id})   "
        f"Designation: {staff.designation or '—'}   "
        f"Shift: {staff.shift_start or '07:00'} – {staff.shift_end or '16:00'}   "
        f"Weekly Off: {staff.weekly_off or 'Sunday'}"
    )
    ws["A2"].font = Font(size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left")

    headers = ["Date", "Day", "Punch In", "Punch Out",
               "Total Hours", "Regular Hours", "OT Hours", "OT Minutes", "Status", "Edited"]
    col_widths = [14, 11, 12, 12, 13, 14, 10, 12, 13, 8]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    totals = {"ot_min": 0, "present": 0, "absent": 0, "partial": 0, "wo": 0}
    last_data_row = 4
    for r_idx, r in enumerate(records, 5):
        last_data_row = r_idx
        try:
            day_name = datetime.strptime(r['date'], "%Y-%m-%d").strftime("%A")
        except Exception:
            day_name = ""
        status = r['status']
        ot_min = r.get('ot_minutes', 0)
        row_data = [
            r['date'], day_name,
            r['punch_in'] if r['punch_in'] != "-" else "",
            r['punch_out'] if r['punch_out'] != "-" else "",
            r['total_hours'], r['regular_hours'], r['ot_hours'], ot_min,
            status, "Yes" if r['is_edited'] else "",
        ]
        fill_color = status_colors.get(status, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.border = thin
            cell.alignment = center
            cell.fill = row_fill
        totals['ot_min'] += ot_min
        if status == "Present":      totals['present'] += 1
        elif status == "Absent":     totals['absent'] += 1
        elif status == "Partial":    totals['partial'] += 1
        elif status == "Weekly Off": totals['wo'] += 1

    summary_row = last_data_row + 2
    total_ot_h, total_ot_m = divmod(totals['ot_min'], 60)
    summary_vals = [\"SUMMARY\", \"\",
                    f\"Present: {totals['present']}\", f\"Absent: {totals['absent']}\",
                    f\"Partial: {totals['partial']}\", f\"Weekly Off: {totals['wo']}\",
                    f\"Total OT: {total_ot_h}h {total_ot_m}m\", totals['ot_min'], \"\", \"\"]
    for col, val in enumerate(summary_vals, 1):
        cell = ws.cell(row=summary_row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = sum_fill
        cell.border = thin
        cell.alignment = center

    ws.freeze_panes = \"A5\"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
\"\"\"

with open(target_file, "a") as f:
    f.write(func_code)
"""

with open("scratch/add_func.py", "w") as f:
    f.write(CodeContent)
