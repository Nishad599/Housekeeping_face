"""
Attendance API routes - punch-in/out, records, muster book, export.
"""
import io
import json
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional

from app.database import get_db
from app.schemas import PunchResponse, AttendanceEditRequest, BulkManualMarkRequest
from app.services.face_service import process_punch_image
from app.services.attendance_service import (
    record_punch, edit_attendance_record, get_muster_book, log_audit
)
from app.models.attendance import AttendanceRecord, AttendancePunch
from app.models.staff import Staff
from app.auth.auth_service import require_role, get_current_user

router = APIRouter(prefix="/api/attendance", tags=["attendance"])


@router.post("/punch", response_model=PunchResponse)
async def punch(
    face_image: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Main punch endpoint.
    Accepts face image → recognizes → records punch IN/OUT.
    """
    image_bytes = await face_image.read()

    # Face recognition pipeline
    staff_id, confidence, error = process_punch_image(image_bytes, db)

    if staff_id is None:
        # Log failed attempt
        log_audit(db, "PUNCH_REJECTED", "attendance", 0, details={
            "error": error, "confidence": confidence
        })
        return PunchResponse(
            success=False,
            message=error,
            confidence=confidence,
        )

    # Record the punch
    success, message, punch_data = record_punch(db, staff_id, confidence)

    return PunchResponse(
        success=success,
        message=message,
        employee_id=punch_data.get("employee_id") if punch_data else None,
        employee_name=punch_data.get("name") if punch_data else None,
        punch_type=punch_data.get("punch_type") if punch_data else None,
        punch_time=datetime.fromisoformat(punch_data["punch_time"]) if punch_data and punch_data.get("punch_time") else None,
        confidence=confidence,
    )


@router.get("/today")
def get_today_attendance(
    db: Session = Depends(get_db),
):
    """Get today's attendance for all staff."""
    today = date.today()
    records = db.query(AttendanceRecord, Staff).join(
        Staff, AttendanceRecord.staff_id == Staff.id
    ).filter(
        AttendanceRecord.date == today
    ).order_by(Staff.employee_id).all()

    return [
        {
            "id": rec.id,
            "employee_id": staff.employee_id,
            "name": staff.name,
            "designation": staff.designation,
            "punch_in": rec.punch_in_time.strftime("%I:%M %p") if rec.punch_in_time else None,
            "punch_out": rec.punch_out_time.strftime("%I:%M %p") if rec.punch_out_time else None,
            "total_hours": f"{rec.total_work_minutes // 60}h {rec.total_work_minutes % 60}m",
            "ot_hours": f"{rec.ot_minutes // 60}h {rec.ot_minutes % 60}m",
            "status": rec.status,
            "is_edited": rec.is_edited,
        }
        for rec, staff in records
    ]


@router.get("/staff-today/{employee_id}")
def get_staff_today_attendance(
    employee_id: str,
    db: Session = Depends(get_db)
):
    """Get today's attendance for a specific staff member (Public access)."""
    today = date.today()
    staff = db.query(Staff).filter(Staff.employee_id == employee_id).first()
    
    if not staff:
        raise HTTPException(status_code=404, detail="Employee not found")
        
    rec = db.query(AttendanceRecord).filter(
        AttendanceRecord.staff_id == staff.id,
        AttendanceRecord.date == today
    ).first()
    
    if not rec:
        return {"found": False, "employee_name": staff.name}
        
    total = rec.total_work_minutes or 0
    return {
        "found": True,
        "employee_name": staff.name,
        "punch_in": rec.punch_in_time.strftime("%I:%M %p") if rec.punch_in_time else None,
        "punch_out": rec.punch_out_time.strftime("%I:%M %p") if rec.punch_out_time else None,
        "total_hours": f"{total // 60}h {total % 60}m",
        "status": rec.status
    }


@router.get("/muster")
def muster_book(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    employee_id: Optional[str] = None,
    name: Optional[str] = None,
    designation: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Get monthly muster book data. Filters: employee_id, name (partial), designation (partial)."""
    return get_muster_book(db, year, month, employee_id, name, designation, location)


@router.get("/muster-matrix")
def muster_matrix(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    employee_id: Optional[str] = None,
    name: Optional[str] = None,
    designation: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor", "viewer")),
):
    """Get monthly muster book in matrix format. Filters: employee_id, name, designation."""
    from app.services.attendance_service import get_muster_matrix
    return get_muster_matrix(db, year, month, employee_id, name, designation, location)


@router.put("/record/{record_id}")
def edit_record(
    record_id: int,
    req: AttendanceEditRequest,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor")),
):
    """Edit an attendance record (admin/supervisor only)."""
    success, message = edit_attendance_record(
        db, record_id,
        req.punch_in_time, req.punch_out_time,
        req.status, req.edit_reason,
        editor=user.full_name,
    )
    if not success:
        raise HTTPException(status_code=400, detail=message)
    return {"message": message}


@router.post("/manual-mark/{employee_id}")
def manual_mark(
    employee_id: str,
    req: AttendanceEditRequest,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor")),
):
    """Manually mark attendance for today (admin/supervisor only)."""
    from app.services.attendance_service import manual_mark_attendance
    success, message = manual_mark_attendance(
        db, employee_id,
        req.punch_in_time, req.punch_out_time,
        req.status, req.edit_reason,
        editor=user.full_name,
    )
    if not success:
        raise HTTPException(status_code=400, detail=message)
    return {"message": message}


@router.post("/bulk-manual-mark")
def bulk_manual_mark(
    req: BulkManualMarkRequest,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor")),
):
    """Bulk mark attendance for multiple staff on a given date (admin/supervisor only)."""
    from app.services.attendance_service import bulk_manual_mark_attendance
    target_date = datetime.strptime(req.date, "%Y-%m-%d").date()
    result = bulk_manual_mark_attendance(
        db,
        employee_ids=req.employee_ids,
        target_date=target_date,
        punch_in_str=req.punch_in_time,
        punch_out_str=req.punch_out_time,
        status=req.status,
        reason=req.edit_reason,
        editor=user.full_name,
    )
    return result


@router.get("/export/muster")
def export_muster_excel(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    employee_id: Optional[str] = None,
    name: Optional[str] = None,
    designation: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor")),
):
    """Export traditional matrix-style muster book as Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    from app.services.attendance_service import get_muster_matrix
    matrix = get_muster_matrix(db, year, month, employee_id, name, designation, location)
    days = matrix["days"]
    staff_data = matrix["staff_data"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Muster Book"

    # Styling colors
    colors = {
        "header": "2C3E50",
        "P": "D5F5E3", # Present - Green
        "A": "FADBD8", # Absent - Red
        "WO": "EBEDEF", # Weekly Off - Gray
        "PL": "FEF9E7"  # Partial - Yellow
    }
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(days) + 8)
    ws["A1"] = f"MUSTER BOOK - {datetime(year, month, 1).strftime('%B %Y').upper()}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Emp ID", "Name"] + [str(d) for d in days] + ["Present", "Absent", "Partial", "W/O", "OT Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data Rows
    for r_idx, s in enumerate(staff_data, 4):
        # Info
        ws.cell(row=r_idx, column=1, value=s["employee_id"]).border = thin_border
        ws.cell(row=r_idx, column=2, value=s["name"]).border = thin_border
        
        # Attendance Days (1-31)
        for c_idx, day in enumerate(days, 3):
            val = s["days"][day]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if val in colors:
                cell.fill = PatternFill(start_color=colors[val], fill_type="solid")
        
        # Summary
        sum_cols = [
            s["summary"]["P"], s["summary"]["A"], s["summary"]["PL"], 
            s["summary"]["WO"], s["summary"]["Total OT"]
        ]
        start_col = len(days) + 3
        for i, val in enumerate(sum_cols):
            cell = ws.cell(row=r_idx, column=start_col + i, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Column Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    for i in range(3, len(days) + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 4
    
    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"muster_matrix_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/individual")
def export_individual_attendance(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    employee_id: str = Query(...),
    db: Session = Depends(get_db),
):
    """Export a single staff member's full monthly attendance as Excel."""
    staff = db.query(Staff).filter(Staff.employee_id == employee_id).first()
    if not staff:
        raise HTTPException(status_code=404, detail=f"Employee '{employee_id}' not found")

    from app.services.attendance_service import get_muster_book, generate_individual_report_excel
    records = get_muster_book(db, year, month, employee_id=employee_id)
    
    buf = generate_individual_report_excel(staff, records, year, month)
    
    filename = f"attendance_{employee_id}_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/bulk-individual")
def export_bulk_individual(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    employee_id: Optional[str] = None,
    name: Optional[str] = None,
    designation: Optional[str] = None,
    location: Optional[str] = None,
    db: Session = Depends(get_db),
    user=Depends(require_role("admin", "supervisor")),
):
    """Export individual reports for all matched staff as a ZIP of Excel files."""
    import zipfile
    import io
    from app.services.attendance_service import get_muster_book, generate_individual_report_excel
    
    # Get staff list based on filters
    query = db.query(Staff).filter(Staff.is_active == True)
    if employee_id:
        query = query.filter(Staff.employee_id == employee_id)
    if name:
        query = query.filter(Staff.name.ilike(f"%{name}%"))
    if designation:
        query = query.filter(Staff.designation.ilike(f"%{designation}%"))
    if location:
        query = query.filter(Staff.location.ilike(f"%{location}%"))
    
    staff_list = query.all()
    if not staff_list:
        raise HTTPException(status_code=404, detail="No staff members found matching filters")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for staff in staff_list:
            records = get_muster_book(db, year, month, employee_id=staff.employee_id)
            if not records: continue
            
            excel_buf = generate_individual_report_excel(staff, records, year, month)
            filename = f"attendance_{staff.employee_id}_{year}_{month:02d}.xlsx"
            zf.writestr(filename, excel_buf.getvalue())

    zip_buf.seek(0)
    zip_filename = f"bulk_attendance_{year}_{month:02d}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_filename}"'},
    )


@router.get("/punches")

def get_punches(
    date_str: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Get raw punch records for debugging/audit."""
    query = db.query(AttendancePunch, Staff).join(
        Staff, AttendancePunch.staff_id == Staff.id
    )

    if date_str:
        punch_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        from sqlalchemy import func
        query = query.filter(func.date(AttendancePunch.punch_time) == punch_date)

    if employee_id:
        query = query.filter(Staff.employee_id == employee_id)

    query = query.order_by(AttendancePunch.punch_time.desc()).limit(100)

    return [
        {
            "id": p.id,
            "employee_id": s.employee_id,
            "name": s.name,
            "punch_type": p.punch_type,
            "punch_time": p.punch_time.isoformat(),
            "confidence": round(p.confidence, 3) if p.confidence else None,
            "is_valid": p.is_valid,
            "rejection_reason": p.rejection_reason,
        }
        for p, s in query.all()
    ]
